#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 工时模板填充模块

读取 Excel 模板，将工时分配数据填入对应字段。
支持的模板字段：任务名称、预计工时、计划开始日期、计划结束日期、任务描述

导出规则（简化后）：
- 工时按权重分配为整数（避免出现小数）
- 小任务按“任务汇总”合并，尽可能保证单条工时 >= 2h
- 任务名称与任务描述均完整列出全部内容（不做截断概括）
"""
from __future__ import annotations

import logging
import math
from copy import copy
from pathlib import Path

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)

# 需要识别并填充的表头关键字 -> 内部字段名
HEADER_FIELD_MAP: dict[str, str] = {
    "任务名称": "task_name",
    "预计工时": "hours",
    "计划开始日期": "start_date",
    "计划结束日期": "end_date",
    "任务描述": "description",
}


def _format_full_list(names: list[str]) -> str:
    """完整列出全部任务名；多条时编号换行。"""
    names = [n.strip() for n in names if (n or "").strip()]
    if not names:
        return ""
    if len(names) == 1:
        return names[0]
    return "\n".join(f"{i}. {n}" for i, n in enumerate(names, 1))


def _check_openpyxl() -> None:
    if not OPENPYXL_AVAILABLE:
        raise ImportError("未安装 openpyxl，请运行: pip install openpyxl")


def _find_header_row(ws) -> int | None:
    """扫描工作表，返回包含表头关键字的行号（1-indexed）。"""
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if cell.value and str(cell.value).strip() in HEADER_FIELD_MAP:
                return row_idx
    return None


def _find_column_map(ws, header_row: int) -> dict[str, int]:
    """返回 {字段名: 列号} 映射。"""
    col_map: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value:
            key = str(cell.value).strip()
            if key in HEADER_FIELD_MAP:
                col_map[HEADER_FIELD_MAP[key]] = cell.column
    return col_map


def _copy_cell_style(src_cell, dst_cell) -> None:
    """将源单元格的样式复制到目标单元格。"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.border = copy(src_cell.border)
        dst_cell.number_format = src_cell.number_format


def _freeze_cell_style(cell) -> dict | None:
    """在删行前冻结样式，避免 delete_rows 后 Cell 引用失效。"""
    if cell is None or not cell.has_style:
        return None
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "alignment": copy(cell.alignment),
        "border": copy(cell.border),
        "number_format": cell.number_format,
    }


def _apply_frozen_style(dst_cell, style: dict | None) -> None:
    if not style:
        return
    dst_cell.font = style["font"]
    dst_cell.fill = style["fill"]
    dst_cell.alignment = style["alignment"]
    dst_cell.border = style["border"]
    dst_cell.number_format = style["number_format"]


def _collect_row_defaults(ws, header_row: int, max_col: int) -> list[tuple]:
    """
    从表头下旧数据行提取每列默认值与样式。
    指派人/确认人/状态等非变更列：取首个非空值；样式取首个有样式的单元格。
    """
    data_start = header_row + 1
    data_end = int(ws.max_row or header_row)
    defaults: list[tuple] = []
    for col_idx in range(1, max_col + 1):
        value = None
        style = None
        for row_idx in range(data_start, data_end + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if style is None:
                style = _freeze_cell_style(cell)
            if value is None and cell.value not in (None, ""):
                value = cell.value
        defaults.append((value, style))
    return defaults


def merge_and_normalize_tasks(tasks: list[dict]) -> list[dict]:
    """
    对任务列表进行合并与规范化（针对同一天的任务）。

    目标：
    - 导出工时为整数（避免 0.30 / 0.11 等小数）
    - 每条输出工时尽可能保证 >= 2h
    - 任务名称与任务描述均完整列出全部内容
    """
    if not tasks:
        return []

    def full_name(item: dict) -> str:
        return str(item.get("task_name", "")).strip()

    # 1) 按权重映射到“整数小时”（largest remainder）
    total_hours = sum(float(t.get("hours", 0) or 0) for t in tasks)
    if total_hours <= 0:
        return []

    target_total_int = int(round(total_hours))
    if target_total_int == 0:
        return []

    ordered = sorted(
        tasks,
        key=lambda t: (float(t.get("hours", 0) or 0), str(t.get("task_name", ""))),
        reverse=True,
    )

    exact_alloc = []
    for t in ordered:
        h = float(t.get("hours", 0) or 0)
        exact_alloc.append(h / total_hours * target_total_int)

    floors = [int(math.floor(v + 1e-9)) for v in exact_alloc]
    fracs = [v - f for v, f in zip(exact_alloc, floors)]
    remainder = target_total_int - sum(floors)

    allocated_int = floors[:]
    if remainder > 0:
        order = sorted(range(len(ordered)), key=lambda i: fracs[i], reverse=True)
        for k in range(remainder):
            allocated_int[order[k % len(order)]] += 1
    elif remainder < 0:
        order = sorted(range(len(ordered)), key=lambda i: fracs[i])
        for k in range(-remainder):
            idx = order[k % len(order)]
            if allocated_int[idx] > 0:
                allocated_int[idx] -= 1

    # 2) 将 <2h 的整数任务归为小任务池，并合并成“汇总条目”
    big_items: list[dict] = []
    small_items: list[dict] = []
    for t, alloc, exact in zip(ordered, allocated_int, exact_alloc):
        alloc_int = int(alloc)
        if alloc_int <= 0:
            continue
        item = {
            **t,
            "allocated_hours": alloc_int,
            "exact_hint": float(exact),
        }
        if alloc_int >= 2:
            big_items.append(item)
        else:
            small_items.append(item)

    big_items.sort(key=lambda x: (x["allocated_hours"], x["exact_hint"]), reverse=True)
    small_items.sort(key=lambda x: (x["allocated_hours"], x["exact_hint"]), reverse=True)

    result: list[dict] = []
    small_total = sum(i["allocated_hours"] for i in small_items)

    # 若小任务合计不足 2h，则并入最高优先级的大任务，避免产生 1h 输出条目
    absorbed_items: list[dict] = []
    if small_items and big_items and small_total < 2:
        absorbed_items = list(small_items)
        big_items[0]["allocated_hours"] += small_total
        small_total = 0

    # 输出大任务：名称与描述均完整列出
    for idx, item in enumerate(big_items):
        name = full_name(item)
        detail_names = [name]
        if idx == 0 and absorbed_items:
            for a in absorbed_items:
                an = full_name(a)
                if an:
                    detail_names.append(an)
        full_text = _format_full_list(detail_names)
        result.append({
            "task_name": full_text,
            "hours": int(item["allocated_hours"]),
            "start_date": item.get("start_date", ""),
            "end_date": item.get("end_date", ""),
            "description": full_text,
        })

    # 输出小任务汇总（仅当小任务合计 >=2）
    if small_items and small_total >= 2:
        merged_names = [full_name(i) for i in small_items if full_name(i)]
        full_text = _format_full_list(merged_names)
        base = small_items[0]
        result.append({
            "task_name": full_text,
            "hours": int(small_total),
            "start_date": base.get("start_date", ""),
            "end_date": base.get("end_date", ""),
            "description": full_text,
        })

    # 兜底：单日总工时理论上为 8，因此输出小时不会超出 8
    for r in result:
        if float(r.get("hours", 0) or 0) > 8.0:
            r["hours"] = 8

    return result


def collect_tasks(
    work_hours_data: dict,
    project_filters: list[str] | None = None,
) -> list[dict]:
    """
    从工时数据中收集任务列表，可按项目名称过滤，并按日期做合并规范化。

    Args:
        work_hours_data: calculate_work_hours() 返回的字典
        project_filters: 项目名称列表（精确匹配，为 None 或空则收集全部）

    Returns:
        list of task dicts, 已做合并规范化（每条 hours ∈ [1, 8]）
    """
    tasks_by_date: dict[str, list[dict]] = {}

    for date_str, date_data in sorted(work_hours_data.items()):
        for project_path, project_data in date_data.get("projects", {}).items():
            project_name = project_data.get("project_name", project_path)

            # 项目过滤（精确匹配项目名）
            if project_filters:
                if project_name not in project_filters:
                    continue

            for task in project_data.get("tasks", []):
                commit_id = task.get("commit_id", "")
                desc_parts = [task["task_name"]]
                if task.get("task_type"):
                    desc_parts.append(f"[{task['task_type']}]")
                if commit_id:
                    desc_parts.append(f"(commit: {commit_id})")

                tasks_by_date.setdefault(date_str, []).append({
                    "task_name": task["task_name"],
                    "hours": task["hours"],
                    "start_date": date_str,
                    "end_date": date_str,
                    "description": " ".join(desc_parts),
                    "project_name": project_name,
                    "task_type": task.get("task_type", ""),
                })

    # 按日期排序，每天内做合并规范化
    result: list[dict] = []
    for date_str in sorted(tasks_by_date.keys()):
        result.extend(merge_and_normalize_tasks(tasks_by_date[date_str]))

    return result


def fill_excel_template(
    template_path: str | Path,
    work_hours_data: dict,
    output_path: str | Path,
    project_filters: list[str] | None = None,
) -> int:
    """
    将工时数据填入 Excel 模板并保存。

    模板约定：
    - 存在一行表头，包含"任务名称"、"预计工时"等关键字
    - 表头下旧行提供样式，以及指派人/确认人/状态等非变更列的默认值
    - 导出前清空旧任务行，再写入本次数据；仅覆盖任务名/工时/日期/描述

    Args:
        template_path: Excel 模板文件路径
        work_hours_data: calculate_work_hours() 返回的工时数据
        output_path: 输出 Excel 文件路径
        project_filters: 要导出的项目名称列表（精确匹配；None 则导出全部）

    Returns:
        写入的任务行数

    Raises:
        ImportError: openpyxl 未安装
        ValueError: 模板格式不符合预期 / 无可用数据
        FileNotFoundError: 模板文件不存在
    """
    _check_openpyxl()

    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    tasks = collect_tasks(work_hours_data, project_filters)
    if not tasks:
        filter_hint = f"（已选项目：{', '.join(project_filters)}）" if project_filters else ""
        raise ValueError(f"未找到任何任务数据{filter_hint}，请先生成工时报告")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 1. 定位表头行
    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError(
            "找不到表头行。请确认 Excel 模板包含以下列之一：\n"
            + "、".join(HEADER_FIELD_MAP.keys())
        )

    # 2. 获取列映射
    col_map = _find_column_map(ws, header_row)
    missing = [k for k, v in HEADER_FIELD_MAP.items() if v not in col_map]
    if missing:
        logger.warning("模板中未找到以下列（将跳过）: %s", "、".join(missing))

    if not col_map:
        raise ValueError("模板中没有可识别的目标列，请检查表头名称")

    # 3. 从旧数据提取默认值/样式（删行前冻结，保留指派人等非变更列）
    example_row_idx = header_row + 1
    max_col = ws.max_column
    fill_cols = set(col_map.values())
    example_row_data = _collect_row_defaults(ws, header_row, max_col)

    # 4. 清空表头下方全部旧任务行（避免上月明细残留）
    old_rows = max(0, int(ws.max_row or header_row) - header_row)
    if old_rows > 0:
        ws.delete_rows(example_row_idx, old_rows)

    # 5. 逐任务插入行
    insert_at = example_row_idx
    for i, task in enumerate(tasks):
        current_row = insert_at + i
        ws.insert_rows(current_row)

        # 非变更列沿用模板旧值；变更列稍后覆盖
        for col_idx, (default_val, style) in enumerate(example_row_data, start=1):
            dst = ws.cell(row=current_row, column=col_idx)
            if col_idx not in fill_cols:
                dst.value = default_val
            _apply_frozen_style(dst, style)

        # 覆盖任务名/工时/日期/描述
        field_values = {
            "task_name": task["task_name"],
            "hours": task["hours"],
            "start_date": task["start_date"],
            "end_date": task["end_date"],
            "description": task["description"],
        }
        for field, col_idx in col_map.items():
            if field in field_values:
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = field_values[field]
                val = field_values[field]
                if field in ("task_name", "description") and isinstance(val, str) and "\n" in val:
                    align = copy(cell.alignment) if cell.alignment else openpyxl.styles.Alignment()
                    align.wrap_text = True
                    align.vertical = "top"
                    cell.alignment = align
                    ws.row_dimensions[current_row].height = max(
                        ws.row_dimensions[current_row].height or 15,
                        15 * (val.count("\n") + 1),
                    )

    wb.save(output_path)
    logger.info("Excel 导出完成：%s（共 %d 行）", output_path, len(tasks))
    return len(tasks)


def list_projects(work_hours_data: dict) -> list[str]:
    """返回工时数据中所有项目名称的去重排序列表。"""
    projects: set[str] = set()
    for date_data in work_hours_data.values():
        for project_data in date_data.get("projects", {}).values():
            projects.add(project_data.get("project_name", ""))
    return sorted(p for p in projects if p)


def parse_work_hours_md(content: str) -> dict:
    """
    将工时分配报告（Markdown 格式）解析回 work_hours_data 字典。

    支持单日和多日报告格式。返回的字典与 calculate_work_hours() 输出兼容。

    Args:
        content: Markdown 文件内容

    Returns:
        work_hours_data dict

    Raises:
        ValueError: 未找到有效的工时数据
    """
    import re

    result: dict = {}

    date_pattern = re.compile(r"\*\*统计日期\*\*[：:]\s*(\d{4}-\d{2}-\d{2})")
    total_hours_pattern = re.compile(r"\*\*标准工时\*\*[：:]\s*([\d.]+)")

    date_matches = list(date_pattern.finditer(content))
    if not date_matches:
        raise ValueError(
            "未找到工时数据。\n"
            "请确认所选文件是「工时分配报告」（包含「**统计日期**」字段）。"
        )

    for i, date_match in enumerate(date_matches):
        date_str = date_match.group(1)
        start = date_match.start()
        end = date_matches[i + 1].start() if i + 1 < len(date_matches) else len(content)
        section = content[start:end]

        th_match = total_hours_pattern.search(section)
        total_hours = float(th_match.group(1)) if th_match else 8.0

        # 解析 Markdown 表格行
        # 新格式: | **项目名** (4.5h) | 任务名 | 任务类型 | 2.50 | commitid | url |
        # 旧格式: | **项目名** (4.5h) | 任务名 | 任务类型 | 2.50 | commitid | branch | url |
        table_row_pat = re.compile(r"^\|(.+)\|$", re.MULTILINE)
        projects: dict = {}
        current_project: str | None = None

        for row_str in table_row_pat.findall(section):
            cells = [c.strip() for c in row_str.split("|")]
            # 需要至少 4 列：项目/空、任务名、任务类型、工时
            if len(cells) < 4:
                continue
            # 跳过表头和分隔行
            if "项目名称" in cells[0] or "------" in cells[0]:
                continue

            project_cell = cells[0]
            task_name = cells[1] if len(cells) > 1 else ""
            task_type = cells[2] if len(cells) > 2 else ""
            hours_str = cells[3] if len(cells) > 3 else "0"
            commit_id = (cells[4] if len(cells) > 4 else "").strip()
            # 兼容有/无「分支」列
            col5 = (cells[5] if len(cells) > 5 else "").strip()
            col6 = (cells[6] if len(cells) > 6 else "").strip()
            if col6 or (col5.startswith("http://") or col5.startswith("https://")):
                branch = "" if (col5.startswith("http://") or col5.startswith("https://")) else col5
            else:
                branch = col5

            # 判断是否有新项目名
            if project_cell:
                proj_match = re.match(r"\*\*(.+?)\*\*\s*\(([\d.]+)h\)", project_cell)
                if proj_match:
                    current_project = proj_match.group(1)
                    project_total = float(proj_match.group(2))
                    if current_project not in projects:
                        projects[current_project] = {
                            "project_name": current_project,
                            "total_hours": project_total,
                            "tasks": [],
                        }

            if not current_project or not task_name:
                continue

            try:
                hours = float(hours_str)
            except ValueError:
                hours = 0.0

            projects[current_project]["tasks"].append({
                "task_name": task_name,
                "task_type": task_type,
                "hours": hours,
                "commit_id": commit_id,
                "branch": branch,
                "commit_url": "",
                "gitlab_url": "",
                "commits": 1,
                "additions": 0,
                "deletions": 0,
            })

        if projects:
            result[date_str] = {
                "date": date_str,
                "total_hours": total_hours,
                "projects": projects,
            }

    if not result:
        raise ValueError("文件中未解析到有效的项目/任务数据，请检查报告格式。")

    return result


def load_work_hours_file(path: str) -> dict:
    """
    自动识别文件类型（.json / .md），加载并返回 work_hours_data 字典。

    Args:
        path: 文件路径（.json 或 .md）

    Returns:
        work_hours_data dict

    Raises:
        ValueError: 格式不支持或解析失败
        FileNotFoundError: 文件不存在
    """
    import json
    from pathlib import Path as _Path

    p = _Path(path)
    if not p.exists():
        raise FileNotFoundError(f"文件不存在: {path}")

    suffix = p.suffix.lower()
    content = p.read_text(encoding="utf-8")

    if suffix == ".json":
        data = json.loads(content)
        if not isinstance(data, dict) or not data:
            raise ValueError("JSON 文件格式不正确，应为工时数据字典")
        return data
    elif suffix in (".md", ".markdown"):
        return parse_work_hours_md(content)
    else:
        # 尝试 JSON 优先，再尝试 MD
        try:
            data = json.loads(content)
            if isinstance(data, dict) and data:
                return data
        except Exception:
            logger.debug("JSON解析工时数据失败，尝试MD格式解析")
        return parse_work_hours_md(content)
