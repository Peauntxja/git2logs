#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据 tests/helpers 中的固定样本重新生成 tests/expected 下的 golden 文件。

用法（仓库根目录）:
  python3 scripts/update_golden_fixtures.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tests.helpers.fake_gitlab import (
    build_sample_all_results,
    build_grouped_commits,
)
from tests.helpers.normalize import normalize_report_text
from report_generator import (
    generate_statistics_report,
    generate_work_hours_report,
    generate_daily_report,
    generate_markdown_log,
)
from work_hours import calculate_work_hours
from excel_exporter import merge_and_normalize_tasks

EXPECTED = ROOT / "tests" / "expected"
AUTHOR = "MIZUKI"
SINCE = "2026-01-15"
UNTIL = "2026-01-15"


def _write(name: str, content: str) -> None:
    path = EXPECTED / name
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    print(f"  wrote {path.relative_to(ROOT)}")


def _fake_commit_details(_project, commit):
    msg = (commit.message or "").split("\n")[0]
    return {
        "short_message": msg,
        "full_message": commit.message or "",
        "stats": None,
        "changed_files": [],
    }


def _build_excel_template(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工时"
    headers = ["任务名称", "预计工时", "计划开始日期", "计划结束日期", "任务描述"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=1, value="示例任务")
    ws.cell(row=2, column=2, value=2)
    ws.cell(row=2, column=3, value="2026-01-01")
    ws.cell(row=2, column=4, value="2026-01-01")
    ws.cell(row=2, column=5, value="示例")
    wb.save(path)


def main() -> int:
    import unittest.mock as mock

    all_results = build_sample_all_results()
    grouped = build_grouped_commits()

    print("Generating golden fixtures...")
    EXPECTED.mkdir(parents=True, exist_ok=True)

    stats = generate_statistics_report(
        all_results, AUTHOR, since_date=SINCE, until_date=UNTIL,
    )
    _write("statistics.md", normalize_report_text(stats))

    with mock.patch("report_generator.get_commit_details", side_effect=_fake_commit_details):
        daily = generate_daily_report(
            all_results, AUTHOR,
            since_date=SINCE, until_date=UNTIL, branch="main",
        )
    _write("daily_report.md", normalize_report_text(daily))

    work_md = generate_work_hours_report(
        all_results, AUTHOR,
        since_date=SINCE, until_date=UNTIL, branch="main",
    )
    _write("work_hours.md", normalize_report_text(work_md))

    commits_md = generate_markdown_log(
        grouped, AUTHOR, repo_name="demo/project-a", project=None,
    )
    _write("commits.md", normalize_report_text(commits_md))

    wh_data = calculate_work_hours(
        all_results, since_date=SINCE, until_date=UNTIL, branch="main",
    )
    _write(
        "work_hours_data.json",
        json.dumps(wh_data, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
    )

    # Excel 合并规则样本（单日任务列表）
    sample_tasks = []
    for _date, date_data in wh_data.items():
        for _proj, pdata in date_data.get("projects", {}).items():
            for task in pdata.get("tasks", []):
                sample_tasks.append({
                    "task_name": task["task_name"],
                    "hours": task["hours"],
                    "start_date": _date,
                    "end_date": _date,
                    "description": task["task_name"],
                    "task_type": task.get("task_type", ""),
                })
    merged = merge_and_normalize_tasks(sample_tasks)
    _write(
        "excel_merged_tasks.json",
        json.dumps(merged, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
    )

    tpl = ROOT / "tests" / "fixtures" / "work_hours_template.xlsx"
    tpl.parent.mkdir(parents=True, exist_ok=True)
    _build_excel_template(tpl)
    print(f"  wrote {tpl.relative_to(ROOT)}")

    meta = {
        "author": AUTHOR,
        "since_date": SINCE,
        "until_date": UNTIL,
        "excel_expected_row_count": len(merged),
    }
    _write("meta.json", json.dumps(meta, ensure_ascii=False, indent=2) + "\n")

    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
